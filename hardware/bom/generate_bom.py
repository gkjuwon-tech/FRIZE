# -*- coding: utf-8 -*-
"""
FRIZE 하드웨어 BOM 생성기
 - FRIZE_BOM.xlsx (요약 + VISOR-1 + SCOUT-1 + 공구/소모품 시트)
 - FRIZE_BOM.pdf  (한글 표 PDF)
가격은 전부 '추정 단가(USD)'다. 실제 구매가는 환율/공급사/MOQ에 따라 출렁인다.
"""
import os
from datetime import date

USD_KRW = 1380  # 환산용 환율 (대충 잡음)

# ---------------------------------------------------------------------
#  BOM 데이터  (부품명, 스펙/모델, 수량, 단가USD, 비고)
# ---------------------------------------------------------------------
VISOR = [
    ("메인 컴퓨트 모듈", "NVIDIA Jetson Orin NX 16GB (엣지 VLM 추론)", 1, 599, "현장 판단을 여기서 직접 돌림"),
    ("컴퓨트 캐리어보드", "Orin NX 커스텀 캐리어 (소형 폼팩터)", 1, 180, ""),
    ("열화상 카메라", "FLIR Boson 640 (640x512, 60Hz)", 1, 3500, "★심장. 연기 뚫는 눈"),
    ("저조도 RGB 카메라", "Sony STARVIS 글로벌셔터 모듈", 1, 220, ""),
    ("AR 디스플레이", "양안 도파관 + 1080p microOLED 모듈", 1, 1200, "시각 명령 출력부"),
    ("IMU", "Bosch BMI088 (산업용 6축)", 1, 25, "머리 자세/움직임"),
    ("RTK GNSS", "u-blox ZED-F9P", 1, 190, "옥외 cm급 위치"),
    ("UWB 실내측위 모듈", "Qorvo DWM3000 (DW3110, 802.15.4z)", 1, 28, "★실내 0.1~0.3m 측위(GPS 차폐 보완)"),
    ("UWB 안테나", "UWB 칩 안테나 + 동축", 1, 8, "nav 포드 모노폴"),
    ("CO 가스센서", "Alphasense 전기화학식 CO", 1, 60, ""),
    ("O2 가스센서", "전기화학식 산소 농도", 1, 80, "산소결핍 경보"),
    ("가연성가스(LEL) 센서", "촉매연소/NDIR LEL", 1, 120, "폭발하한 감지"),
    ("H2S/HCN 센서", "전기화학식 황화수소/시안화수소", 1, 90, "유독가스"),
    ("온도 측정보드", "K타입 열전대 + 주변온도", 1, 35, ""),
    ("MEMS 마이크 어레이", "3채널 빔포밍", 1, 18, "음성→텍스트"),
    ("골전도 트랜스듀서", "양쪽 골전도 스피커", 2, 20, "귀 막혀도 들림"),
    ("5G 모뎀", "Quectel RM502Q-AE", 1, 260, ""),
    ("서브기가 메시 라디오", "RFD900x 계열 mesh", 1, 130, "통신 두절 대비"),
    ("WiFi6E/BT M.2", "근거리 고대역", 1, 35, ""),
    ("안테나 세트", "5G/GNSS/메시/WiFi", 1, 60, ""),
    ("핫스왑 배터리팩", "21700 듀얼셀 내열팩", 2, 120, "현장에서 갈아끼움"),
    ("전원관리 보드(PMIC)", "커스텀 다중레일 PMIC", 1, 90, ""),
    ("냉각 어셈블리", "히트파이프 + 마이크로 블로워(밀폐)", 1, 55, ""),
    ("AR 커버 렌즈", "내열/내스크래치 곡면 커버", 1, 90, ""),
    ("외부 하우징", "CNC 알루미늄 + 세라믹 코팅 (프로토)", 1, 250, "★타협X"),
    ("내부 프레임", "PC-CF 프린트 부품 일체", 1, 40, "scad에서 뽑음"),
    ("FFC/FPC 플렉스케이블", "카메라/디스플레이 배선 세트", 1, 45, ""),
    ("체결/커넥터/배선", "방진방수 커넥터, 하네스", 1, 70, ""),
    ("스토리지", "NVMe 256GB", 1, 40, "로그/모델"),
    ("헬멧/SCBA 마운트", "NFPA 헬멧 레일 + 마스크 인터페이스", 1, 80, ""),
]

SCOUT = [
    ("센터 플레이트(쌍)", "카본 4T CNC 상/하판", 1, 160, "scad center_plate"),
    ("카본 암 튜브", "16mm 외경 카본튜브", 6, 15, ""),
    ("암 클램프/모터마운트", "PA-CF 프린트", 6, 10, "scad arm_clamp"),
    ("브러시리스 모터", "T-Motor 고추력 (예: MN605S 급)", 6, 90, "★추력 여유"),
    ("ESC", "고전류 6-in-1 또는 개별", 1, 260, ""),
    ("프로펠러(쌍)", "카본 폴딩 프롭", 6, 20, ""),
    ("비행 컨트롤러", "Pixhawk 6X / Cube Orange+", 1, 480, "이중화 IMU"),
    ("온보드 컴퓨트", "Jetson Orin NX 16GB", 1, 599, "기내 VLM"),
    ("컴퓨트 캐리어", "Orin 캐리어보드", 1, 160, ""),
    ("LiDAR", "Livox Mid-360 (연기 속 SLAM)", 1, 749, "★연기 뚫고 맵핑"),
    ("열화상(짐벌)", "FLIR Boson 640", 1, 3500, "★"),
    ("RGB 짐벌 카메라", "저조도 디지털 카메라", 1, 300, ""),
    ("3축 짐벌", "내열 브러시리스 짐벌", 1, 350, ""),
    ("RTK GNSS", "u-blox ZED-F9P", 1, 190, ""),
    ("텔레메트리/메시", "장거리 mesh 라디오", 1, 130, ""),
    ("디지털 영상 송신", "저지연 디지털 VTX", 1, 200, ""),
    ("배터리", "6S 16000mAh 고방전 LiPo", 2, 130, "비행시간 확보"),
    ("전원분배보드(PDB)", "고전류 PDB + XT90", 1, 60, ""),
    ("가스 센서 포드", "CO/LEL/O2 통합 포드", 1, 260, ""),
    ("세라믹 열차폐 키트", "세라믹 블랭킷 + 반사막", 1, 180, "★복사열 차단"),
    ("랜딩기어", "카본 랜딩기어 세트", 1, 70, ""),
    ("방진댐퍼/스탠드오프", "짐벌/FC 방진 + 하드웨어", 1, 80, ""),
    ("배선/커넥터", "실리콘 와이어, XT90, 방수커넥터", 1, 70, ""),
    ("스토리지", "NVMe 256GB", 1, 40, ""),
    ("UWB 앵커 디스펜서", "3발 매거진 + 서보 게이트(scad DISPENSER)", 1, 60, "★진입 시 비콘 투하"),
    ("디스펜서 서보+릴리즈", "메탈기어 서보 + 게이트 슬라이드", 1, 25, "1발씩 낙하"),
]

TOOLS = [
    ("납땜 스테이션", "온도조절 (Hakko 급)", 1, 250, "차고 필수"),
    ("열풍 리워크", "SMD 작업용", 1, 120, ""),
    ("멀티미터", "트루RMS", 1, 60, ""),
    ("스마트 LiPo 충전기", "밸런스 충전", 1, 200, "★불나면 안되니 좋은거"),
    ("토크 드라이버 세트", "정밀 토크", 1, 90, "드론 볼트 토크관리"),
    ("크림핑 툴 세트", "커넥터 압착", 1, 70, ""),
    ("디지털 캘리퍼스", "0.01mm", 1, 40, ""),
    ("ESD 매트+스트랩", "정전기 방지", 1, 40, "Jetson 보호"),
    ("열수축/솔더/플럭스", "소모품", 1, 50, ""),
    ("써멀 페이스트/패드", "방열", 1, 30, ""),
    ("나사풀림방지(Loctite)", "243 중강도", 1, 15, "드론 진동 대비"),
    ("3D 프린터(없으면)", "엔지니어링 필라멘트 대응", 1, 700, "선택 / PA-CF 출력"),
]

# ---------------------------------------------------------------------
#  SW ↔ HW 연결/인터페이스 (소프트웨어가 하드웨어를 실제로 구동하는 접점)
#  FRIZE 소프트웨어(Jetson의 frize_visor/frize_scout)가 센서·FC·LiDAR·카메라와
#  물리적으로 통신하기 위한 브리지 부품. 이게 있어야 SW가 HW를 '딸깍' 움직인다.
# ---------------------------------------------------------------------
INTERFACE = [
    ("센서 통합 MCU 보드", "STM32F4 ― 가스4종/IMU 집약→Jetson(USB-CDC), frize_visor sensors HAL", 2, 18, "고글+가스포드 펌웨어"),
    ("USB-UART 어댑터", "FTDI FT232H ― Pixhawk TELEM2 ↔ Jetson(MAVLink/MAVSDK)", 2, 12, "frize_scout flight"),
    ("GMSL2 SerDes 보드", "열화상/RGB 카메라 장거리 시리얼라이저-디시리얼라이저", 3, 120, "FLIR/Sony→Jetson CSI"),
    ("LiDAR 이더넷 어댑터+스위치", "Livox Mid-360 기가비트 인터페이스 + 소형 산업 스위치", 1, 70, "frize_mapping 입력"),
    ("레벨 시프터/버스 보드", "3.3↔5V, I2C/SPI 분기(센서 다중 연결)", 2, 20, ""),
    ("ST-Link V3 디버거", "MCU 펌웨어 플래싱/디버그", 1, 30, "센서보드 부트로더"),
    ("CAN 트랜시버 모듈", "ESC/배터리 텔레메트리(드론 상태→SW)", 2, 15, ""),
    ("커넥터/시그널 케이블 키트", "JST-GH, Molex, 실드 신호선(전 모듈 배선)", 1, 45, ""),
    ("microSD 128GB(산업용)", "Jetson OS/FRIZE 이미지 플래싱", 3, 20, "고글2+드론1"),
    ("절연 USB 허브", "UVC 카메라/주변장치 다중 + 노이즈 절연", 2, 35, ""),
    ("RTC/PPS 시각동기 모듈", "센서·영상 타임스탬프 정합(융합 정확도)", 1, 25, "frize_mapping 정합"),
    ("PoE 인젝터/DC-DC", "컴퓨트·LiDAR 안정 급전(전원↔데이터)", 1, 40, ""),
]


# ---------------------------------------------------------------------
#  FRIZE CONSOLE-1 ― 전용 현장 지휘 콘솔 (노트북 대체, 1 지휘소당 1대)
#  러기드 케이스 + 대형 터치스크린 + 물리 컨트롤덱 + E-STOP + 통신 + 배터리.
#  여기서 콕핏(frize_cockpit) SW가 키오스크로 구동된다.
# ---------------------------------------------------------------------
CONSOLE = [
    ("임베디드 컴퓨트", "NVIDIA Jetson AGX Orin 64GB (콕핏+맵+VLM 클라이언트)", 1, 1999, "콘솔 두뇌"),
    ("메인 디스플레이", "21.5\" 1000nit 선샤인리더블 정전터치 산업패널", 1, 900, "★현장 시인성"),
    ("보조 상태 디스플레이", "7\" 바형 IPS(연결/경보 상시표시)", 1, 120, ""),
    ("러기드 인클로저", "IP65 알루미늄+폴리머 케이스(CNC 커스텀)", 1, 400, "★낙하/방수"),
    ("컨트롤덱 버튼", "기계식 일루미네이티드 버튼 16ea(유닛/뷰/기능)", 1, 160, ""),
    ("로터리 인코더+조그다이얼", "광학식 인코더 2 + 조그 1(맵 줌/선택)", 1, 90, ""),
    ("E-STOP 스위치", "긴급중단(전 대원 후퇴 브로드캐스트 트리거)", 1, 25, "★안전"),
    ("사운드/경보", "스피커 + 경보 부저", 1, 30, ""),
    ("전원/UPS 보드", "와이드레인지 DC-DC + 무정전 절체", 1, 120, ""),
    ("핫스왑 배터리", "현장 운용 배터리팩", 2, 120, "교대 운용"),
    ("통신 모듈 세트", "5G + 서브기가 메시 + WiFi6E + 안테나 4", 1, 450, "다중링크 폴백"),
    ("내장 이더넷 스위치", "기가비트(LiDAR/장비 유선)", 1, 60, ""),
    ("냉각 어셈블리", "히트파이프 + 저소음 블로워", 1, 70, ""),
    ("케이블/커넥터", "방수 글랜드, 하네스", 1, 90, ""),
    ("핸들/범퍼/스트랩", "캐링 + 코너 보호", 1, 60, ""),
    ("스토리지", "NVMe 512GB", 1, 60, "OS/로그/맵"),
    ("주변광+GNSS", "조도센서(자동밝기) + 위치", 1, 80, ""),
]


# ---------------------------------------------------------------------
#  FRIZE VENT-1 ― IoT 자동 배연/진입 포트 (콕핏 OPEN 딸깍 → 액추에이터가 개방)
# ---------------------------------------------------------------------
VENT = [
    ("컨트롤 MCU", "ESP32-S3 (메시/WiFi + IO, frize_apparatus 펌웨어)", 1, 12, ""),
    ("메시 라디오", "서브기가 LoRa 모듈(통신두절 대비)", 1, 35, "fleet 메시"),
    ("리니어 액추에이터", "12V 고추력 1500N IP65", 1, 90, "★개방 구동"),
    ("모터 드라이버", "H브리지(BTS7960급) + 전류감지", 1, 18, ""),
    ("리미트/위치 센서", "리미트 스위치 2 + 홀 위치센서", 1, 15, "개방도 피드백"),
    ("전자식 홀드/잠금", "전자석 릴리즈(평시 고정)", 1, 25, ""),
    ("배터리+BMS", "리튬팩 + 보호회로", 1, 60, ""),
    ("전원보드", "DC-DC + 역전압/과전류 보호", 1, 25, ""),
    ("온도/가스 센서", "K열전대 + CO 전기화학", 1, 40, ""),
    ("러기드 인클로저", "IP65 케이스(CNC/프린트)", 1, 120, "scad frize_vent"),
    ("도어프레임 클램프", "후면 마운트 클램프 2", 1, 45, ""),
    ("푸시패드+러버", "개방 접촉 패드", 1, 20, ""),
    ("안테나/LED/부저", "상태표시 + 경보", 1, 20, ""),
    ("배선/방수커넥터", "하네스, IP67 커넥터", 1, 30, ""),
]

# ANCHOR-1: 드론이 투하하는 UWB 측위 비콘(소모성, 다수 비치). 단가×수량.
ANCHOR = [
    ("UWB 비콘 MCU", "ESP32-C3 + Qorvo DWM3000(앵커 모드)", 6, 22, "삼변측량 앵커"),
    ("UWB 안테나", "칩 안테나 + 매칭", 6, 4, ""),
    ("배터리(단셀)", "21700 + 보호회로(저온 대응)", 6, 9, "수 시간 가동"),
    ("자기복원 무게추", "텅스텐/스틸 베이스(굴러서 똑바로)", 6, 6, "self-righting"),
    ("충격흡수 범퍼+셸", "TPU 범퍼 + 고시인성 PC 셸", 6, 8, "scad frize_anchor"),
    ("상태 LED/스위치", "투하·fix 표시, 자동 전원", 6, 3, ""),
]


def subtotal(rows):
    return sum(q*p for _,_,q,p,_ in rows)

# ---------------------------------------------------------------------
#  XLSX 생성
# ---------------------------------------------------------------------
def build_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="C0392B")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=16, color="C0392B")
    money_fmt = '#,##0'
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sheet_from(ws, title, rows):
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A2"] = "단가/합계는 추정치(USD). KRW는 환율 %d 적용." % USD_KRW
        ws["A2"].font = Font(italic=True, size=9, color="888888")
        headers = ["부품", "스펙 / 모델", "수량", "단가(USD)", "소계(USD)", "소계(KRW)", "비고"]
        r0 = 4
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r0, c, h)
            cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        r = r0 + 1
        for name, spec, qty, price, note in rows:
            line = qty*price
            vals = [name, spec, qty, price, line, line*USD_KRW, note]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = border
                if c in (4,5,6): cell.number_format = money_fmt
                if c == 3: cell.alignment = Alignment(horizontal="center")
            r += 1
        # 합계행
        tot = subtotal(rows)
        ws.cell(r, 4, "합계").font = Font(bold=True)
        ws.cell(r, 4).alignment = Alignment(horizontal="right")
        c5 = ws.cell(r, 5, tot); c5.font = Font(bold=True); c5.number_format = money_fmt
        c6 = ws.cell(r, 6, tot*USD_KRW); c6.font = Font(bold=True); c6.number_format = money_fmt
        for c in range(4,7): ws.cell(r,c).fill = PatternFill("solid", fgColor="FDEBD0")
        # 폭
        widths = [22, 40, 7, 12, 13, 15, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return tot

    # Summary
    ws = wb.active; ws.title = "요약"
    v = subtotal(VISOR); s = subtotal(SCOUT); t = subtotal(TOOLS); ifc = subtotal(INTERFACE)
    cn = subtotal(CONSOLE); vn = subtotal(VENT); an = subtotal(ANCHOR)
    ws["A1"] = "FRIZE 하드웨어 BOM 요약"; ws["A1"].font = title_font
    ws["A2"] = "작성일 %s · 단가 추정 USD · 환율 %d KRW/USD" % (date.today().isoformat(), USD_KRW)
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    rows = [
        ("CONSOLE-1 지휘 콘솔 (1대)", cn),
        ("VISOR-1 스마트고글 (1대)", v),
        ("SCOUT-1 정찰드론 (1대)", s),
        ("VENT-1 IoT 배연/진입 포트 (1대)", vn),
        ("ANCHOR-1 UWB 측위 비콘 (6발 세트)", an),
        ("SW↔HW 연결/인터페이스 (1세트)", ifc),
        ("공구/소모품 (1세트, 1회성)", t),
        ("── 지휘소 1식 (콘솔+고글+드론+VENT+앵커+연결+공구)", cn+v+s+vn+an+ifc+t),
        ("참고: 콘솔1+고글3+드론2+VENT2+앵커12+연결+공구1 운영세트", cn + v*3 + s*2 + vn*2 + an*2 + ifc + t),
    ]
    hdr = ["항목", "USD", "KRW"]
    for c,h in enumerate(hdr,1):
        cell = ws.cell(4,c,h); cell.fill=head_fill; cell.font=head_font
        cell.alignment=Alignment(horizontal="center"); cell.border=border
    for i,(label,amt) in enumerate(rows, start=5):
        ws.cell(i,1,label).border=border
        a=ws.cell(i,2,amt); a.number_format=money_fmt; a.border=border
        b=ws.cell(i,3,amt*USD_KRW); b.number_format=money_fmt; b.border=border
        if label.startswith("──"):
            for c in range(1,4): ws.cell(i,c).font=Font(bold=True); ws.cell(i,c).fill=PatternFill("solid",fgColor="FDEBD0")
    ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=14; ws.column_dimensions["C"].width=18

    sheet_from(wb.create_sheet("CONSOLE-1 콘솔"), "FRIZE CONSOLE-1 ― 전용 현장 지휘 콘솔 BOM", CONSOLE)
    sheet_from(wb.create_sheet("VISOR-1 고글"), "FRIZE VISOR-1 ― 스마트 고글 BOM", VISOR)
    sheet_from(wb.create_sheet("SCOUT-1 드론"), "FRIZE SCOUT-1 ― 정찰 드론 BOM", SCOUT)
    sheet_from(wb.create_sheet("VENT-1 IoT포트"), "FRIZE VENT-1 ― IoT 자동 배연/진입 포트 BOM", VENT)
    sheet_from(wb.create_sheet("ANCHOR-1 UWB비콘"), "FRIZE ANCHOR-1 ― 드론 투하 UWB 측위 비콘 BOM (6발)", ANCHOR)
    sheet_from(wb.create_sheet("SW-HW 연결"), "SW ↔ HW 연결 / 인터페이스 (소프트웨어가 하드웨어를 구동하는 접점)", INTERFACE)
    sheet_from(wb.create_sheet("공구·소모품"), "공구 / 소모품 (차고 셋업)", TOOLS)

    wb.save(path)
    print("XLSX 저장:", path)
    print("  CONSOLE=%d  VISOR=%d  SCOUT=%d  VENT=%d  ANCHOR=%d  IFACE=%d  TOOLS=%d USD" % (cn,v,s,vn,subtotal(ANCHOR),ifc,t))

# ---------------------------------------------------------------------
#  PDF 생성
# ---------------------------------------------------------------------
def build_pdf(path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
    pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
    KO = "HYSMyeongJo-Medium"; KOB = "HYGothic-Medium"

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontName=KOB, fontSize=20, textColor=colors.HexColor("#C0392B"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName=KOB, fontSize=13, textColor=colors.HexColor("#C0392B"))
    body = ParagraphStyle("body", parent=styles["Normal"], fontName=KO, fontSize=8.5, leading=11)
    small = ParagraphStyle("small", parent=styles["Normal"], fontName=KO, fontSize=8, textColor=colors.grey)
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontName=KO, fontSize=7.5, leading=9)
    cellb = ParagraphStyle("cellb", parent=styles["Normal"], fontName=KOB, fontSize=7.5, leading=9)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    story = []

    def P(t, st=cell): return Paragraph(str(t).replace("&","&amp;").replace("<","&lt;"), st)

    def make_table(title, rows):
        story.append(Paragraph(title, h2))
        story.append(Spacer(1, 3))
        data = [[P("부품",cellb),P("스펙 / 모델",cellb),P("수량",cellb),
                 P("단가$",cellb),P("소계$",cellb),P("비고",cellb)]]
        for name,spec,qty,price,note in rows:
            data.append([P(name),P(spec),P(qty),P(f"{price:,}"),P(f"{qty*price:,}"),P(note)])
        tot = subtotal(rows)
        data.append([P("",cell),P("",cell),P("",cell),P("합계",cellb),P(f"{tot:,}",cellb),P(f"≈{tot*USD_KRW:,}원",cell)])
        tbl = Table(data, colWidths=[34*mm,58*mm,11*mm,16*mm,18*mm,35*mm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#C0392B")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#BBBBBB")),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("ALIGN",(2,1),(4,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,colors.HexColor("#FBF2F0")]),
            ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#FDEBD0")),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8))

    story.append(Paragraph("🔥 FRIZE 하드웨어 BOM", h1))
    story.append(Paragraph("Firefighting Real-time Integrated Zone-control Environment", small))
    story.append(Paragraph("작성일 %s · 단가 추정치(USD) · 환율 %d KRW/USD · 소방용 = 스펙 타협 없음"
                           % (date.today().isoformat(), USD_KRW), small))
    story.append(Spacer(1, 10))

    v=subtotal(VISOR); s=subtotal(SCOUT); t=subtotal(TOOLS); ifc=subtotal(INTERFACE)
    cn=subtotal(CONSOLE); vn=subtotal(VENT); an=subtotal(ANCHOR)
    total=cn+v+s+vn+an+ifc+t
    sumdata = [[P("항목",cellb),P("USD",cellb),P("KRW",cellb)],
               [P("CONSOLE-1 지휘 콘솔 (1대)"),P(f"{cn:,}"),P(f"{cn*USD_KRW:,}")],
               [P("VISOR-1 고글 (1대)"),P(f"{v:,}"),P(f"{v*USD_KRW:,}")],
               [P("SCOUT-1 드론 (1대)"),P(f"{s:,}"),P(f"{s*USD_KRW:,}")],
               [P("VENT-1 IoT 포트 (1대)"),P(f"{vn:,}"),P(f"{vn*USD_KRW:,}")],
               [P("ANCHOR-1 UWB 비콘 (6발)"),P(f"{an:,}"),P(f"{an*USD_KRW:,}")],
               [P("SW↔HW 연결/인터페이스 (1세트)"),P(f"{ifc:,}"),P(f"{ifc*USD_KRW:,}")],
               [P("공구/소모품 (1세트)"),P(f"{t:,}"),P(f"{t*USD_KRW:,}")],
               [P("지휘소 1식 합계",cellb),P(f"{total:,}",cellb),P(f"{total*USD_KRW:,}",cellb)]]
    st = Table(sumdata, colWidths=[70*mm,40*mm,50*mm])
    st.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#C0392B")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),0.4,colors.grey),
        ("ALIGN",(1,0),(2,-1),"RIGHT"),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#FDEBD0")),
    ]))
    story.append(st); story.append(Spacer(1,12))
    story.append(PageBreak())

    make_table("CONSOLE-1 ― 전용 현장 지휘 콘솔", CONSOLE)
    story.append(PageBreak())
    make_table("VISOR-1 ― 스마트 고글", VISOR)
    story.append(PageBreak())
    make_table("SCOUT-1 ― 정찰 드론", SCOUT)
    story.append(PageBreak())
    make_table("VENT-1 ― IoT 자동 배연/진입 포트", VENT)
    make_table("ANCHOR-1 ― 드론 투하 UWB 측위 비콘 (6발)", ANCHOR)
    make_table("SW ↔ HW 연결 / 인터페이스 (소프트웨어가 하드웨어를 구동하는 접점)", INTERFACE)
    make_table("공구 / 소모품 (차고 셋업)", TOOLS)
    story.append(Paragraph("※ 가격은 공급사/MOQ/환율에 따라 변동. 핵심 부품(열화상·LiDAR·하우징)은 스펙 우선, 가격 후순위.", small))

    doc.build(story)
    print("PDF 저장:", path)

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    build_xlsx(os.path.join(here, "FRIZE_BOM.xlsx"))
    build_pdf(os.path.join(here, "FRIZE_BOM.pdf"))
    v=subtotal(VISOR); s=subtotal(SCOUT); t=subtotal(TOOLS); ifc=subtotal(INTERFACE); cn=subtotal(CONSOLE); vn=subtotal(VENT); an=subtotal(ANCHOR)
    tot=cn+v+s+vn+ifc+t
    print("=== 지휘소 1식 합계: $%d  (approx %d KRW) ===" % (tot, tot*USD_KRW))
